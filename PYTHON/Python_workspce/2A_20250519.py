import openpyxl

def test1():
    """구구단 출력하여 엑셀파일에 저장하기
    """
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("구구단")
    cols = "ABCDEFGH"
    dan = 2
    for col in cols:
        ws[f"{col}1"] = f"{dan}단"
        for i in range(1,10):
            ws[f"{col}{i+1}"] = f"{dan} x {i} = {dan*i}"
        dan += 1   
    
    wb.save("gugudan.xlsx")
    print("구구단 파일이 저장되었습니다!")

# test1()


def test2():
    """엑셀 파일의 내용 읽기
    """
    wb = openpyxl.load_workbook("gugudan.xlsx")
    ws = wb["구구단"]
    
    cols = "ABCDEFGH"
    for col in cols:
        for i in range(1,11):
            print(ws[f'{col}{i}'].value)
        print()

# test2()


class SoccerPlayer(object):
    def __init__(self, name, position, back_number):
        self.name = name
        self.position = position
        self.back_number = back_number
    def change_back_number(self, new_number):
        self.back_number = new_number
    def __str__(self):
        return f"이름은 {self.name}, 포지션은 {self.position}, 등번호는 {self.back_number}"

def test3():
    p1 = SoccerPlayer("홍길동","MF",17)
    print(f"이름은 {p1.name}, 포지션은 {p1.position}, 등번호는 {p1.back_number}")
    print(p1.__str__())
    print(p1)

# test3()


class Circle():
    """원을 나타내는 클래스
    필드: 반지름(radius)
    메소드: 둘레(get_around), 면적(get_area) 구하기
    """
    def __init__(self, radius):
        self.radius = radius
    def get_around(self):
        return 2 * 3.141592 * self.radius
    def get_area(self):
        return 3.141592 * self.radius ** 2
    def __str__(self):
        return f"반지름은 {self.radius}, 둘레는 {self.get_around()}, 면적은 {self.get_area()}"

def test4():
    c1 = Circle(10)
    c2 = Circle(20)
    print(c1)
    print(c2)

# test4()


class Rectangle():
    """사각형 클래스
    속성: 넓이, 높이
    메소드: 면적, 둘레 구하기
    """
    def __init__(self, width, height):
        self.width = width
        self.height = height
    def get_area(self):
        return self.width * self.height
    def get_around(self):
        return (self.width + self.height) * 2
    def __str__(self):
        return f"넓이는 {self.width}, 높이는 {self.height}, 둘레는 {self.get_around()}, 면적은 {self.get_area()}"

def test5():
    r1 = Rectangle(10,5)
    r2 = Rectangle(10,10)
    print(r1)
    print(r2)

# test5()


# 클래스 변수 vs 인스턴스 변수
class Person():
    species = "Homo sapiens"
    def __init__(self, name, age):
        self.name = name
        self.age = age
    def __str__(self):
        return f"이름은 {self.name}, 나이는 {self.age}세, 종은 {Person.species}"

def test6():
    p1 = Person("홍길동",10)
    p2 = Person("홍길순",13)
    print(p1)
    print(p2)

# test6()


# ******************************
# 메소드 오버로딩(X) -> 가변인자
# ******************************
class Calculator():
    """2개의 수에 대해서 사칙연산을 수행하는 계산기
    """
    def __init__(self, num1, num2):
        self.num1 = num1
        self.num2 = num2
    def add(self):
        return self.num1 + self.num2
    def add(self, num1, num2):
        return num1 + num2
    def sub(self):
        return self.num1 - self.num2
    def sub(self, num1, num2):
        return num1 - num2
    def mul(self):
        return self.num1 * self.num2
    def mul(self, num1, num2):
        return num1 * num2
    def div(self):
        return self.num1 / self.num2
    def div(self, num1, num2):
        return num1 / num2

class Calculator2():
    def __init__(self, *args):
        # if len(args) > 0:
        if len(args):
            self.num1 = args[0]
            self.num2 = args[1]
    def add(self, *args):
        if len(args):
            return args[0] + args[1]
        else:
            return self.num1 + self.num2
    def sub(self, *args):
        if len(args):
            return args[0] - args[1]
        else:
            return self.num1 - self.num2
    def mul(self, *args):
        if len(args):
            return args[0] * args[1]
        else:
            return self.num1 * self.num2
    def div(self, *args):
        if len(args):
            return args[0] / args[1]
        else:
            return self.num1 / self.num2

def test7():
    num1 = 10
    num2 = 5
    c1 = Calculator2(num1, num2)
    print("{} + {} = {}".format(num1,num2,c1.add()))
    print("{} - {} = {}".format(num1,num2,c1.sub()))
    print("{} x {} = {}".format(num1,num2,c1.mul()))
    print("{} % {} = {}".format(num1,num2,c1.div()))
    
    num1 = 20
    num2 = 10
    c2 = Calculator2()
    print("{} + {} = {}".format(num1,num2,c2.add(num1,num2)))
    print("{} - {} = {}".format(num1,num2,c2.sub(num1,num2)))
    print("{} x {} = {}".format(num1,num2,c2.mul(num1,num2)))
    print("{} % {} = {}".format(num1,num2,c2.div(num1,num2)))
    
# test7()


class Person():
    def __init__(self, name, age, gender):
        self.name = name
        self.age = age
        self.gender = gender
    def do_work(self):
        print("먹는다!")
    def __str__(self):
        return f"이름은 {self.name}, 나이는 {self.age}, 성별은 {self.gender}"

class Employee(Person):
    def __init__(self, name, age, gender, salary, hire_date):
        super().__init__(name, age, gender)
        self.salary = salary
        self.hire_date = hire_date
    def do_work(self):
        print("열심히 일한다!")
    def __str__(self):
        return f"{super().__str__()}, 봉급은 {self.salary}, 고용일은 {self.hire_date}"

def work(p):
    p.do_work()

def test8():
    p1 = Person("김수지", 21, "여성")
    e1 = Employee("홍길동", 23, "남성", 2500000, "2025-11-02")
    work(p1)
    work(e1)

test8()












