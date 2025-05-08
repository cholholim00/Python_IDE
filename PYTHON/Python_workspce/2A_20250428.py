import openpyxl
from openpyxl import Workbook
import os

# 폴더가 없으면 생성
folder_path = 'EXCEL'
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

def test1():
    # 액셀 만들기
    wb = openpyxl.Workbook()

    # 액셀 시트 만들기
    ws = wb.create_sheet(title='회원정보')

    # 액셀 시트에 데이터 입력하기
    ws['A1'] = '이름'
    ws['B1'] = '전화번호'
    ws['A2'] = '홍길동'
    ws['B2'] = '010-1234-5678'
    ws['A3'] = '이순신'
    ws['B3'] = '010-9876-5432'

    # 액셀 파일 저장하기
    wb.save(os.path.join(folder_path, 'member.xlsx'))
    # # 액셀 파일 저장하기
    # wb.save('member.xlsx')
# test1()

def test2():
    # 액셀 파일 열기
    wb = openpyxl.load_workbook('EXCEL/member.xlsx')

    # 액셀 시트 선택하기
    ws = wb['회원정보']

    # 액셀 시트에서 데이터 바꾸기
    print(ws['A2'])
    ws['A2'] = '김철수'
    wb.save(os.path.join(folder_path,'memberchange.xlsx'))

    # # 액셀 시트에서 데이터 읽기
    # for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
    #     for cell in row:
    #         print(cell.value, end=' ')
    #     print()
# test2()

def test3():
    """
    엑셀 파일에 출력하기
    출력형식
    이름    나이
    홍길동1    20
    홍길동2    30
    ...
    홍길동10   100
    """
    # 액셀 만들기
    wb = openpyxl.Workbook()

    # 액셀 시트 만들기
    ws = wb.create_sheet(title='회원정보')

    # 액셀 시트에 데이터 입력하기
    ws['A1'] = '이름'
    ws['B1'] = '나이'
    for i in range(1, 11):
        ws['A' + str(i + 1)] = '홍길동' + str(i)
        ws['B' + str(i + 1)] = 9 + i

    # 액셀 파일 저장하기
    wb.save(os.path.join(folder_path,'memberauto.xlsx'))
# test3()

def test4(): # 이름 과 나이를 입력받아 액셀에 저장하기

    # 액셀 만들기
    wb = openpyxl.Workbook()

    # 액셀 시트 만들기
    ws = wb.create_sheet(title='회원정보')

    # 액셀 시트에 데이터 입력하기
    ws['A1'] = '이름'
    ws['B1'] = '나이'
    for i in range(1, 11):
        ws['A' + str(i + 1)] = str(input(' 이름 :'))
        ws['B' + str(i + 1)] = str(input(' 나이 :'))

    # 액셀 파일 저장하기
    wb.save(os.path.join(folder_path,'memberinput.xlsx'))
# test4()

